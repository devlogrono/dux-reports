from __future__ import annotations

import os
import re
import unicodedata
from collections import Counter
from copy import deepcopy
from typing import Any

from dux.common.physical.registration import (
    NUMERIC_FIELDS,
    PERIMETER_FIELDS,
    SKINFOLD_FIELDS,
)

from openpyxl import load_workbook


MAX_ISAK_FILE_SIZE_BYTES = 200 * 1024 * 1024


ISAK_FIELD_DECIMALS = {
    "peso_bruto_kg": 1,
    "talla_corporal_cm": 1,
    "talla_sentado_cm": 1,
    "envergadura_cm": 1,
    "acromial_radial": 1,
    "radial_estiloidea": 2,
    "medial_estiloidea_dactilar": 2,
    "ilioespinal": 2,
    "trocanterea": 1,
    "troc_tibial_lateral": 1,
    "tibial_lateral": 1,
    "tibial_medial_maleolar_medial": 2,
    "pie": 2,
    "biacromial": 1,
    "bi_iliocrestideo": 1,
    "humeral_biepicondilar": 1,
    "femoral_biepicondilar": 1,
    "muneca_biestiloideo": 2,
    "tobillo_bimaleolar": 2,
    "torax_transverso": 2,
    "torax_antero_posterior": 2,
    "mano": 2,
    **{field: 1 for field in PERIMETER_FIELDS},
    **{field: 1 for field in SKINFOLD_FIELDS},
}

ISAK_EXCEL_FIELDS = set(NUMERIC_FIELDS)
_FIELD_ALIASES = {
    "longitud_acromial_radial": "acromial_radial",
    "longitud_radial_estiloidea": "radial_estiloidea",
    "longitud_medial_estiloidea_dactilar": "medial_estiloidea_dactilar",
    "longitud_ilioespinal": "ilioespinal",
    "longitud_trocanterea": "trocanterea",
    "longitud_troc_tibial_lateral": "troc_tibial_lateral",
    "longitud_tibial_lateral": "tibial_lateral",
    "longitud_tibial_medial_maleolar_medial": "tibial_medial_maleolar_medial",
    "longitud_pie": "pie",
    "diametro_biacromial": "biacromial",
    "diametro_bi_iliocrestideo": "bi_iliocrestideo",
    "diametro_humeral_biepicondilar": "humeral_biepicondilar",
    "diametro_femoral_biepicondilar": "femoral_biepicondilar",
    "diametro_muneca_biestiloideo": "muneca_biestiloideo",
    "diametro_tobillo_bimaleolar": "tobillo_bimaleolar",
    "diametro_torax_transverso": "torax_transverso",
    "diametro_torax_antero_posterior": "torax_antero_posterior",
    "diametro_mano": "mano",
}


def normalize_isak_excel_key(value: Any, duplicate_counter: Counter | None = None) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")

    units = ""
    match = re.search(r"\((.*?)\)", text)
    if match:
        units = re.sub(r"[^a-z0-9]", "", match.group(1).strip())

    text = re.sub(r"\(.*?\)", "", text)
    text = text.replace("-", "_").replace("/", "_")
    text = re.sub(r"[^a-z0-9_ ]", "", text)
    text = re.sub(r"_+", "_", text.replace(" ", "_")).strip("_")
    if units:
        text = f"{text}_{units}"
    text = re.sub(r"_+", "_", text).strip("_")

    candidates = [text]
    if text.endswith("_kg") or text.endswith("_cm") or text.endswith("_mm"):
        candidates.append(re.sub(r"_(kg|cm|mm)$", "", text))

    for candidate in candidates:
        alias = _FIELD_ALIASES.get(candidate)
        if alias:
            return alias
        if candidate in ISAK_EXCEL_FIELDS:
            return candidate

    base = candidates[-1] if candidates else text
    counter = duplicate_counter if duplicate_counter is not None else Counter()
    count = counter[base]
    counter[base] += 1

    if count == 0:
        candidate = f"perimetro_{base}"
        if candidate in ISAK_EXCEL_FIELDS:
            return candidate
    if count == 1:
        candidate = f"pliegue_{base}"
        if candidate in ISAK_EXCEL_FIELDS:
            return candidate

    for prefix in ("perimetro_", "pliegue_"):
        candidate = f"{prefix}{base}"
        if candidate in ISAK_EXCEL_FIELDS:
            return candidate

    return base


def _excel_filename(file) -> str:
    filename = getattr(file, "filename", None) or getattr(file, "name", "") or ""
    return os.path.basename(filename)


def _file_size(file) -> int | None:
    content_length = getattr(file, "content_length", None)
    if content_length:
        return int(content_length)

    stream = getattr(file, "stream", file)
    try:
        current = stream.tell()
        stream.seek(0, os.SEEK_END)
        size = stream.tell()
        stream.seek(current)
        return size
    except (AttributeError, OSError):
        return content_length


def _validate_excel_file(file) -> str:
    size = _file_size(file)
    if size is not None and size > MAX_ISAK_FILE_SIZE_BYTES:
        raise ValueError("El archivo supera el maximo permitido de 200 MB.")

    filename = _excel_filename(file)
    ext = os.path.splitext(filename)[1].lower()
    if ext not in {".xls", ".xlsx"}:
        raise ValueError("Formato no soportado. Usa un archivo .xls o .xlsx.")
    return ext


def _read_xlsx(file, sheet_name: str | None) -> dict[str, Any]:
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("El archivo no es un Excel .xlsx valido.") from exc

    sheets = workbook.sheetnames
    selected_sheet = sheet_name.strip() if isinstance(sheet_name, str) and sheet_name.strip() else None
    if selected_sheet is None:
        selected_sheet = workbook.active.title
    if selected_sheet not in sheets:
        raise ValueError(f"La hoja '{selected_sheet}' no existe en el archivo.")

    worksheet = workbook[selected_sheet]
    duplicate_counter: Counter = Counter()
    rows = []
    for row_index in range(5, 56):
        raw_key = worksheet.cell(row=row_index, column=1).value
        raw_value = worksheet.cell(row=row_index, column=2).value
        if raw_key is None and raw_value is None:
            continue
        if raw_value is None or str(raw_value).strip() == "":
            continue
        normalized_key = normalize_isak_excel_key(raw_key, duplicate_counter)
        if not normalized_key:
            continue
        rows.append(
            {
                "row": row_index,
                "label": str(raw_key or "").strip(),
                "campo": normalized_key,
                "valor": raw_value,
            }
        )

    workbook.close()
    if not rows:
        raise ValueError("La hoja ISAK no contiene valores utilizables.")

    return {
        "sheets": sheets,
        "selected_sheet": selected_sheet,
        "rows": rows,
    }


def _read_xls(file, sheet_name: str | None) -> dict[str, Any]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("Falta la dependencia xlrd para procesar archivos .xls.") from exc

    stream = getattr(file, "stream", file)
    try:
        stream.seek(0)
        content = stream.read()
        workbook = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise ValueError("El archivo no es un Excel .xls valido.") from exc
    finally:
        try:
            stream.seek(0)
        except (AttributeError, OSError):
            pass

    sheets = workbook.sheet_names()
    selected_sheet = sheet_name.strip() if isinstance(sheet_name, str) and sheet_name.strip() else None
    if selected_sheet is None:
        selected_sheet = sheets[0] if sheets else None
    if selected_sheet not in sheets:
        raise ValueError(f"La hoja '{selected_sheet}' no existe en el archivo.")

    worksheet = workbook.sheet_by_name(selected_sheet)
    duplicate_counter: Counter = Counter()
    rows = []
    for row_index in range(5, 56):
        zero_based = row_index - 1
        if zero_based >= worksheet.nrows:
            continue
        raw_key = worksheet.cell_value(zero_based, 0) if worksheet.ncols > 0 else None
        raw_value = worksheet.cell_value(zero_based, 1) if worksheet.ncols > 1 else None
        if raw_key in (None, "") and raw_value in (None, ""):
            continue
        if raw_value is None or str(raw_value).strip() == "":
            continue
        normalized_key = normalize_isak_excel_key(raw_key, duplicate_counter)
        if not normalized_key:
            continue
        rows.append(
            {
                "row": row_index,
                "label": str(raw_key or "").strip(),
                "campo": normalized_key,
                "valor": raw_value,
            }
        )

    if not rows:
        raise ValueError("La hoja ISAK no contiene valores utilizables.")

    return {
        "sheets": sheets,
        "selected_sheet": selected_sheet,
        "rows": rows,
    }


def read_isak_excel(file, sheet_name: str | None = None) -> dict[str, Any]:
    ext = _validate_excel_file(file)
    if ext == ".xls":
        return _read_xls(file, sheet_name)
    return _read_xlsx(file, sheet_name)


def read_isak_xlsx(file, sheet_name: str | None = None) -> dict[str, Any]:
    return read_isak_excel(file, sheet_name=sheet_name)


def analyze_isak_excel_fields(rows: list[dict[str, Any]], min_fields: int = 40) -> dict[str, Any]:
    excel_fields = {str(row.get("campo")) for row in rows if row.get("campo")}
    matched = excel_fields & ISAK_EXCEL_FIELDS
    missing = ISAK_EXCEL_FIELDS - excel_fields
    coverage_pct = round((len(matched) / len(ISAK_EXCEL_FIELDS)) * 100, 2) if ISAK_EXCEL_FIELDS else 0

    return {
        "mapper": {field: field for field in matched},
        "matched": sorted(matched),
        "missing": sorted(missing),
        "coverage_pct": coverage_pct,
        "is_valid": len(matched) >= min_fields,
        "matched_count": len(matched),
        "total_fields": len(ISAK_EXCEL_FIELDS),
    }


def _parse_excel_float(value: Any, decimals: int) -> float | None:
    if value is None:
        return None
    try:
        number = float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None
    return round(number, decimals)


def _record_values_from_rows(rows: list[dict[str, Any]], mapper: dict[str, str]) -> dict[str, Any]:
    record = {}
    for row in rows:
        campo = row.get("campo")
        if campo not in mapper:
            continue
        field = mapper[campo]
        decimals = ISAK_FIELD_DECIMALS.get(field, 1)
        record[field] = _parse_excel_float(row.get("valor"), decimals)
    return record


def build_registration_record_from_file(
    file,
    sheet_name: str | None,
    base_record: dict[str, Any],
    current_user,
) -> dict[str, Any]:
    del current_user
    parsed = read_isak_excel(file, sheet_name=sheet_name)
    analysis = analyze_isak_excel_fields(parsed["rows"])
    record = deepcopy(base_record)
    record.update(_record_values_from_rows(parsed["rows"], analysis["mapper"]))
    record["tipo_isak"] = "COMPLETO"
    record["metodo"] = "ISAK"

    return {
        "record": record,
        "analysis": analysis,
        "rows": parsed["rows"],
        "sheets": parsed["sheets"],
        "selected_sheet": parsed["selected_sheet"],
    }
